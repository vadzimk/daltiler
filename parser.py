import shutil

import openpyxl
import csv
import tabula
from openpyxl import Workbook
import subprocess

import TEMPLATE
import functions as f  # contains major functions
import PCONST as PC  # contains pdf constants

import sys  # access to functions and variables that allow for working with Python interpreter
import os

# do refactor to make a row a custom object with all the methods

args = sys.argv  # get the list of arguments

infilename = args[1] if len(args) == 2 else 'test1.pdf'  # infilename is the args[1] or default to 'test1.pdf'

print(f"Working on {infilename}\nPlease wait....")
midfilename = 'out.csv'
outfilename = 'output.xlsx'




# -------------------- UNCOMMENT THIS TO CONVERT TO HTML -----------------
# # cleanup directory
# if os.path.exists("xpdf") and os.path.isdir("xpdf"):
#     print("xpdf exists, deleting it....")
#     shutil.rmtree("xpdf")
#f.convert_pdf_to_html('test1.pdf')
# -------------------------------------------------------------------

currow = 1  # contains the current row in the output file

# wb = openpyxl.load_workbook(filename) # excel file
# sheet = wb[wb.sheetnames[0]]
wb = Workbook()  # create excel workbook
sheet = wb.active  # get active sheet of the workbook

""" create template header """

# complete row 1 - header
for i in range(len(TEMPLATE.HEADER)):
    sheet.cell(row=currow, column=i + 1, value=TEMPLATE.HEADER[i])

# read pdf into data frame
# df = tabula.read_pdf('test5.pdf', stream=True, pages='all')
# print(df)

# convert pdf into csv
# ------------------------- UNCOMMENT THIS CONVERT PDF -----------------------------
tabula.convert_into(infilename, midfilename, output_format='csv', pages='all', stream=True)
# ---------------------------------------------------------

# read the csv file call it csvfile
with open(midfilename, newline='') as csvfile:
    readerObject = csv.reader(csvfile, dialect='excel')  # returns reader object that is an iterator
    listOfRows = list(readerObject)

    # fields of pdf  # will be processed first
    series_name = ""
    group_name = ""
    subgroup_name = ""
    vendor_code = ""
    item_size = ""
    item_color = ""
    item_units_per_carton = ""
    units_of_measure = ""
    unit_price = ""

    # fields of template  # will be populated after manual review
    item_name = ""
    display_name = ""
    item_number_and_name = ""

    for r in range(len(listOfRows)):
        row_obj = listOfRows[r]  # row_obj is a of class list

        if f.contains_series(row_obj):
            series_name = f.get_series_name(row_obj)

        if f.contains_group(row_obj):
            group_name = row_obj[0]

        if f.contains_subgroup(row_obj):
            SUBGROUP_INDEX = 2
            subgroup_name = row_obj[SUBGROUP_INDEX]

        if f.contains_vendor_code(row_obj):
            ITEM_SIZE_INDEX = 0
            vendor_code = f.get_vendor_code(row_obj)

            item_size = "".join(row_obj[ITEM_SIZE_INDEX].split()[0:3])

        # print(len(row_obj), row_obj)  # test printout row

        if f.is_valid_row(row_obj):
            """this pdf_line is a valid table row"""
            currow += 1  # go to the next row of the outfile to process
            externalid = "{}-{:05d}".format(PC.VENDOR_NAME_CODE, (currow - 1))  # formatted string
            sheet.cell(row=currow, column=1, value=externalid)
            item_name = series_name + " " + group_name + " " + subgroup_name
            sheet.cell(row=currow, column=4, value=item_name)
            sheet.cell(row=currow, column=6, value=vendor_code)
            sheet.cell(row=currow, column=11, value=item_size)
            item_color = f.get_item_color(row_obj)
            sheet.cell(row=currow, column=10, value=item_color)
            sheet.cell(row=currow, column=9, value=f.get_units_per_carton(row_obj))
            sheet.cell(row=currow, column=7, value=f.get_units_of_measure(row_obj))  # compound field
            sheet.cell(row=currow, column=27, value=f.get_unit_price(row_obj))

            # populate columns derived from data extracted from pdf only
            sheet.cell(row=currow, column=2, value=externalid)
            display_name = item_name + " " + item_size + " " + item_color + " " + vendor_code
            display_name = " ".join(display_name.split())  # remove multiple spaces
            sheet.cell(row=currow, column=3, value=display_name)
            item_number_and_name = externalid + " " + item_name
            sheet.cell(row=currow, column=5, value=item_number_and_name)

# ---------------- UNCOMMENT THIS SAVE OUTFILE ------------------------
wb.save(outfilename)
# -------------------------------------------------------------------
wb.close()

print("Done!")  # indicate the end of the program
