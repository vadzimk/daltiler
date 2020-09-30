import openpyxl
import os
wb = openpyxl.load_workbook('example.xlsx')

print(type(wb))

print(os.getcwd())

print('sheetnames:')
print(wb.sheetnames)
sheet = wb['Sheet3']
print(sheet.title)
anotherSheet = wb.active
print(anotherSheet)

b1 = anotherSheet['B1']  # returns a Cell object
# B1 is a cell's coordinate attribute
print(f'b1 is a ${b1}')
print(b1.value)
print('Row %s, Column %s is %s' % (b1.row, b1.column, b1.value))

print('Cell %s is %s' % (b1.coordinate, b1.value))

# using the sheet's cell() method to pass integers for row and column arguments
# the first row or column integer is 1 , not 0

c = anotherSheet.cell(row=1, column=2)
print(c)
print(c.value)

for i in range(1,8,2): # go through every other row
    print(i, anotherSheet.cell(row=i, column=2).value)