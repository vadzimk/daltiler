import openpyxl
import csv
import tabula
from openpyxl import Workbook
infilename = 'test5.pdf'
outfilename = 'output.xlsx'
midfilename = 'out.csv'

# wb = openpyxl.load_workbook(filename) # excel file
# sheet = wb[wb.sheetnames[0]]
wb = Workbook()
sheet = wb.active


# read pdf into data frame
# df = tabula.read_pdf('test5.pdf', stream=True, pages='all')
# print(df)

# convert pdf into csv
tabula.convert_into(infilename, midfilename, output_format='csv', pages='all', stream=True)


with open(midfilename, newline='') as csvfile:
    spamreader = csv.reader(csvfile, delimiter=',', quotechar='|')
    rows = list(spamreader)
    for r in range(len(rows)):
        print(rows[r])

        for i in range(len(rows[r])):
            sheet.cell(row=r+1, column=i + 1, value=rows[r][i])


wb.save(outfilename)
wb.close()
